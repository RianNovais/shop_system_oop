from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from datetime import datetime
from _email import Email

# This class aims to take the data from the lists that exist in the "shop" class that contain customers, products and
# sales and export them to XLSX individually through specific methods, and that's what we see here, at the beginning
# of the class we pass the paths to 3 files, respectively .xlsx for products, customers and sales and the export date,
# then we have functions that do this work, where they receive the list of products, sales and customers and export them
# to the defined path of the spreadsheet.

e = Email()
class Excel():
    def __init__(self):
        self.pathproductspreadsheet = Path().absolute() / 'spreadsheets' / 'products.xlsx'
        self.pathcustomerspreadsheet = Path().absolute() / 'spreadsheets' / 'customers.xlsx'
        self.pathsalespreadsheet = Path().absolute() / 'spreadsheets' / 'sales.xlsx'
        self.updateDate = datetime.now().strftime('%d/%m/%Y:%H/%M')

    def export_products_to_xlsx(self, listProducts):
        #creating workbook and create an sheet and remove "Sheet" default

        workbook = Workbook()
        workbook.create_sheet('Products')
        del workbook['Sheet']
        #workbook.active is sheet "Products"
        worksheet: Worksheet = workbook.active

        #add headers
        worksheet.cell(1, 1, 'Id')
        worksheet.cell(1, 2, 'Add Date')
        worksheet.cell(1, 3, 'Name')
        worksheet.cell(1, 4, 'Manufacturer')
        worksheet.cell(1, 5, 'Serial Number')
        worksheet.cell(1, 6, 'Category')
        worksheet.cell(1, 7, 'Quantity')
        worksheet.cell(1, 8, 'Price')

        # for each product it generates a list with the attribute values and adds each item respectively
        # on each line with .append (it already recognizes that our 1 line has our titles)
        for product in listProducts:
            prod = [product.id, product.addDate, product.name, product.manufacturer, product.serialNumber,
                    product.category.value, product.quantity, product.price]
            worksheet.append(prod)

        #save and showing success message
        workbook.save(self.pathproductspreadsheet)
        print(f'{worksheet.max_row -1} lines of products add sucessfully to products.xlsx')

        choiceEmail = input('Send spreadsheet in email: [Y/N]: ')
        if choiceEmail.lower() == 'y':
            e.sendProductSpreadsheetEmail(self.pathproductspreadsheet)
    def export_customers_to_xlsx(self, listCustomers):
        workbook = Workbook()
        workbook.create_sheet('Customers')
        del workbook['Sheet']
        worksheet: Worksheet = workbook.active

        # defining headers in a simplified way, passing a list with the titles we want
        # pro worksheet.append, in this case it will fill our first line with each title in each
        # column, forming our header
        headers = ['Id', 'Add Date', 'Name', 'LastName', 'Document', 'Address']
        worksheet.append(headers)



        for customer in listCustomers:
            cust = [customer.id, customer.addDate, customer.name, customer.lastName, customer.document, customer.address]
            worksheet.append(cust)

        workbook.save(self.pathcustomerspreadsheet)
        print(f'{worksheet.max_column -1} lines of customers add sucessfully to customers.xlsx')

        choiceEmail = input('Send spreadsheet in email: [Y/N]: ')
        if choiceEmail.lower() == 'y':
            e.sendCustomerSpreadsheetEmail(self.pathcustomerspreadsheet)
    def export_sales_to_xlsx(self, listSales):
        workbook = Workbook()
        workbook.create_sheet('Sales')
        del workbook['Sheet']

        worksheet: Worksheet = workbook.active


        #add headers
        headers = ['Sale Date', 'Customer', 'Product', 'Quantity']
        worksheet.append(headers)

        for sale in listSales:
            sal = [sale.saleDate, sale.customer.name, sale.product.name, sale.quantity]
            worksheet.append(sal)


        workbook.save(self.pathsalespreadsheet)
        print(f'{worksheet.max_row -1} lines of sales add sucessfully to products.xlsx')

        choiceEmail = input('Send spreadsheet in email: [Y/N]: ')
        if choiceEmail.lower() == 'y':
            e.sendSaleSpreadsheetEmail(self.pathsalespreadsheet)





if __name__ == "__main__":
    ...